"""
🔧 НАСТРОЙКА ПОД ТЕМУ (ЗАМЕНИТЬ ПЕРЕМЕННЫЕ ПОД СВОЮ ПРЕДМЕТНУЮ ОБЛАСТЬ)
================================================================================
MAPPING_CONFIG - Настройка маппинга колонок из файла в БД

Примеры для разных тем:

=== КАФЕ/РЕСТОРАН ===
- orders: колонка A -> title (Название заказа), B -> description (Описание)
- users: колонка A -> login, B -> full_name, C -> role, D -> password_hash
- items: колонка A -> order_id, B -> name (Блюдо), C -> quantity, D -> price

=== АВТОСЕРВИС ===
- orders: колонка A -> title (Номер наряда), B -> description (Неисправность)
- users: колонка A -> login, B -> full_name, C -> role (diagnost/mechanic)
- items: колонка A -> order_id, B -> name (Запчасть), C -> quantity, D -> price

=== ПОЛИКЛИНИКА ===
- orders: колонка A -> title (Номер талона), B -> description (Жалобы)
- users: колонка A -> login, B -> full_name, C -> role (doctor/nurse)
- items: колонка A -> order_id, B -> name (Услуга), C -> quantity, D -> price
================================================================================
"""

import argparse
import json
import sys
from datetime import datetime

try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("⚠️ openpyxl не установлен. Excel файлы не поддерживаются.")
    print("   Установите: pip install openpyxl")

# Подключаемся к моделям Flask
# Важно: нужно инициализировать Flask app перед использованием
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app import app, db
from models import User, Order, Item, Reference

# ==============================================================================
# 🔧 КОНФИГУРАЦИЯ МАППИНГА КОЛОНОК
# ==============================================================================

MAPPING_CONFIG = {
    # Маппинг для пользователей (users)
    'users': {
        'json': {
            'login': 'login',
            'full_name': 'full_name',
            'role': 'role',
            'password_hash': 'password_hash'
        },
        'excel': {
            'A': 'login',
            'B': 'full_name',
            'C': 'role',
            'D': 'password_hash'
        },
        'defaults': {
            'is_blocked': False,
            'password_hash': 'scrypt:32768:8:1$...'  # Нужно указать хеш или сгенерировать
        }
    },
    
    # Маппинг для заказов (orders)
    'orders': {
        'json': {
            'title': 'title',
            'description': 'description',
            'status': 'status',
            'author_id': 'author_id'
        },
        'excel': {
            'A': 'title',
            'B': 'description',
            'C': 'status',
            'D': 'author_id'
        },
        'defaults': {
            'status': 'new',
            'date_created': datetime.utcnow
        }
    },
    
    # Маппинг для позиций заказа (items)
    'items': {
        'json': {
            'order_id': 'order_id',
            'name': 'name',
            'quantity': 'quantity',
            'price': 'price'
        },
        'excel': {
            'A': 'order_id',
            'B': 'name',
            'C': 'quantity',
            'D': 'price'
        },
        'defaults': {
            'quantity': 1,
            'price': 0.0
        }
    },
    
    # Маппинг для справочника (references)
    'references': {
        'json': {
            'name': 'name',
            'type': 'type',
            'price': 'price'
        },
        'excel': {
            'A': 'name',
            'B': 'type',
            'C': 'price'
        },
        'defaults': {
            'price': 0.0
        }
    }
}


def column_letter_to_index(letter):
    """Преобразование буквы колонки Excel в индекс (A->0, B->1, ...)"""
    return ord(letter.upper()) - ord('A')


def load_json_file(filepath):
    """Загрузка данных из JSON файла"""
    with open(filepath, 'r', encoding='utf-8') as f:
        return json.load(f)


def load_excel_file(filepath, sheet_name=None):
    """Загрузка данных из Excel файла"""
    if not EXCEL_SUPPORT:
        raise ImportError("openpyxl не установлен")
    
    wb = load_workbook(filepath, read_only=True)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    data = []
    headers = None
    
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = row
            continue
        data.append(row)
    
    return data, headers


def import_data(file_path, data_type, preview=False):
    """
    Универсальная функция импорта данных
    
    Args:
        file_path: Путь к файлу (JSON или XLSX)
        data_type: Тип данных (users, orders, items, references)
        preview: Если True, только показать данные без сохранения
    """
    
    if data_type not in MAPPING_CONFIG:
        print(f"❌ Неизвестный тип данных: {data_type}")
        print(f"   Доступные типы: {', '.join(MAPPING_CONFIG.keys())}")
        return False
    
    config = MAPPING_CONFIG[data_type]
    
    # Определяем тип файла
    is_excel = file_path.endswith(('.xlsx', '.xls'))
    is_json = file_path.endswith('.json')
    
    if not is_excel and not is_json:
        print("❌ Неизвестный формат файла. Используйте .json или .xlsx")
        return False
    
    # Загружаем данные
    try:
        if is_json:
            raw_data = load_json_file(file_path)
            # Если это список словарей
            if isinstance(raw_data, list):
                records = raw_data
            else:
                records = [raw_data]
        else:
            rows, headers = load_excel_file(file_path)
            records = []
            mapping = config.get('excel', {})
            
            for row in rows:
                record = {}
                for col_letter, field_name in mapping.items():
                    col_idx = column_letter_to_index(col_letter)
                    if col_idx < len(row):
                        record[field_name] = row[col_idx]
                records.append(record)
    except Exception as e:
        print(f"❌ Ошибка чтения файла: {e}")
        return False
    
    print(f"\n{'=' * 60}")
    print(f"📄 Файл: {file_path}")
    print(f"📊 Тип данных: {data_type}")
    print(f"📝 Найдено записей: {len(records)}")
    print(f"{'=' * 60}\n")
    
    if preview:
        print("👀 РЕЖИМ ПРЕДПРОСМОТРА (данные не сохраняются)\n")
    
    # Обработка записей
    saved_count = 0
    skipped_count = 0
    error_count = 0
    
    mapping = config.get('json', {}) if is_json else config.get('excel', {})
    defaults = config.get('defaults', {})
    
    # Определяем модель
    model_map = {
        'users': User,
        'orders': Order,
        'items': Item,
        'references': Reference
    }
    model = model_map[data_type]
    
    for i, record in enumerate(records[:10], 1):  # Показываем первые 10 для preview
        print(f"Запись #{i}:")
        
        # Преобразуем данные согласно маппингу
        processed = {}
        
        for src_field, dest_field in mapping.items():
            value = record.get(src_field) or record.get(dest_field)
            if value is not None:
                processed[dest_field] = value
        
        # Добавляем значения по умолчанию
        for field, default_value in defaults.items():
            if field not in processed:
                if callable(default_value):
                    processed[field] = default_value()
                else:
                    processed[field] = default_value
        
        # Выводим обработанные данные
        for key, value in processed.items():
            print(f"  {key}: {value}")
        
        # Сохраняем если не preview
        if not preview:
            try:
                # Проверка на дубликаты (для users по login)
                if data_type == 'users' and 'login' in processed:
                    existing = User.query.filter_by(login=processed['login']).first()
                    if existing:
                        print(f"  ⚠️ Пропущено: пользователь {processed['login']} уже существует")
                        skipped_count += 1
                        continue
                
                obj = model(**processed)
                db.session.add(obj)
                saved_count += 1
                print(f"  ✅ Добавлено")
                
            except Exception as e:
                print(f"  ❌ Ошибка: {e}")
                error_count += 1
        
        print()
    
    if len(records) > 10:
        print(f"... и ещё {len(records) - 10} записей\n")
    
    # Коммитим транзакцию
    if not preview and saved_count > 0:
        try:
            db.session.commit()
            print(f"{'=' * 60}")
            print(f"✅ Импортировано: {saved_count} записей")
            if skipped_count > 0:
                print(f"⚠️ Пропущено (дубликаты): {skipped_count}")
            if error_count > 0:
                print(f"❌ Ошибок: {error_count}")
            print(f"{'=' * 60}")
        except Exception as e:
            db.session.rollback()
            print(f"❌ Ошибка при сохранении: {e}")
            return False
    
    return True


def main():
    parser = argparse.ArgumentParser(
        description='Универсальный скрипт импорта данных для Flask-приложения'
    )
    parser.add_argument('--file', '-f', required=True, help='Путь к файлу (JSON или XLSX)')
    parser.add_argument('--type', '-t', required=True, 
                       choices=['users', 'orders', 'items', 'references'],
                       help='Тип импортируемых данных')
    parser.add_argument('--preview', '-p', action='store_true',
                       help='Режим предпросмотра (без сохранения в БД)')
    
    args = parser.parse_args()
    
    with app.app_context():
        success = import_data(args.file, args.type, args.preview)
        sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
